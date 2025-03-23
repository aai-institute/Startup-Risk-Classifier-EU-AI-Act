# Standard Library
import os
import re
import ast
import json

# Third-Party Library
import pandas as pd
import openpyxl
from docx import Document
from dotenv import load_dotenv
from openai import OpenAI
import anthropic

# Local Imports
from Classes import ChatGPT, Prompts, WebScraper, TextExtractor
from large_prompts.master_prompt import master_prompt


# Load environment variables
load_dotenv()

# Constants
TOTAL_PAGE_CRAWLS = 4

def claude_api(prompt):
    try:
        client = anthropic.Anthropic(
            api_key=os.getenv("ANTHROPIC_KEY")
        )
        message = client.messages.create(
            model="claude-3-7-sonnet-20250219",
            max_tokens=8192,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        message_content = message.content[0].text  # Extracts the actual message text
        input_tokens = message.usage.input_tokens  # Extracts input tokens
        output_tokens = message.usage.output_tokens  # Extracts output tokens

    except Exception as e:
        message_content = f"API ERROR: Anthropic API failed"
        input_tokens = 0
        output_tokens = 0

    return message_content, input_tokens, output_tokens

def load_startups_excel(startups_file):
    sheet = openpyxl.load_workbook(startups_file)["AI Use Cases"]
    return sheet

def create_results_file():    
    # Save the results to a new Excel file
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "AI Use Cases"

    return output_sheet, output_wb

def content_shortener(content_shortener_model, prompts_obj, page_content):
    chat_shorten_page_obj = ChatGPT(content_shortener_model, prompts_obj.shorten_page_content(page_content), [], OpenAI(api_key=os.getenv("MY_KEY"), max_retries=5))
    chat_shorten_page_response, input_tokens, output_tokens = chat_shorten_page_obj.chat_model()
    # print(f"Shortened Page Content: {chat_shorten_page_response}")

    return chat_shorten_page_response, input_tokens, output_tokens

def read_docx(file_path):
    doc = Document(file_path)
    text = "\n".join([paragraph.text for paragraph in doc.paragraphs]) 
    return text

def prepare_AI_Act_prompt(master_prompt, all_ai_use_cases):

    additional_format = f"""\nDo not give any intros or outros. Respond in plain text string only (no unusual arrays), without any formatting f.e. no bold or ## headings or numbered headings. The following are the AI Use cases of the startup you have to classify:\n\n{all_ai_use_cases}"""

    # print(file_content)  # Output all content
    return master_prompt + additional_format

def traverse_links(web_scraper_obj, links, model_name, content_shortener_model, ai_use_cases, prompts_obj):
    
    try:
        # Traverse the important links
        for link in links:
            web_scraper_obj.set_url(link)
            # print(f"Going into relavant link: {link}")
            web_scraper_obj.load_page()
            page_content = web_scraper_obj.get_page_content(model_name)

            # Shorten the content
            shortened_content, input_tokens, output_tokens = content_shortener(content_shortener_model, prompts_obj, page_content)
            # Update token cost
            web_scraper_obj.set_token_cost(input_tokens, output_tokens, content_shortener_model)


            chat_use_case_obj = ChatGPT(model_name, prompts_obj.update_startup_summary(f"\n\n".join(ai_use_cases), shortened_content), [], OpenAI(api_key=os.getenv("MY_KEY"), max_retries=5))
            chat_use_case_response, input_tokens, output_tokens = chat_use_case_obj.chat_model()

            # Update token cost
            web_scraper_obj.set_token_cost(input_tokens, output_tokens, model_name)
            
            ai_use_cases.append(chat_use_case_response)
    
    except Exception as e:
        # Return error message only
        return ["Page Error: Unable to traverse links"]


    return ai_use_cases

# Extract Python-style list from a string
def extract_list(input_string):
    # Regular expression to match Python-style lists
    list_pattern = r"\[.*?\]"
    match = re.search(list_pattern, input_string)
    if match:
        return ast.literal_eval(match.group())
    return None


def save_to_excel(output_sheet, output_wb, startup_name, url, redirect_url, use_cases_combined, eu_ai_act_response, total_token_cost, output_filename):
    headers = ["Startup Name", "Homepage URL", "Redirected URL", "AI Use Cases" , "EU AI Act Risk Classification", "Total Token Cost ($)"]
    
    # Write headers if not present
    if output_sheet.max_row < 2:
        output_sheet.append(headers)

    # Write data
    
    row = [startup_name, url, redirect_url, use_cases_combined, eu_ai_act_response, total_token_cost]

    output_sheet.append(row)
    output_wb.save(f"{output_filename}")


def extract_use_cases_from_response(use_cases_full_text):
    all_use_cases = []
    # Extract each AI use case from the full text
    text_extractor = TextExtractor(use_cases_full_text)
    text_extractor.set_use_cases()
    df_use_cases = text_extractor.get_use_cases()
    
    for index, row in df_use_cases.iterrows():
        row_string = "\n".join(f"{col}: {val}" for col, val in row.items())
        all_use_cases.append(row_string + '\n\n')
    
    return all_use_cases    



def prompt_approach(classification_model_name, sheet, output_sheet, output_wb, output_filename):
    # Initialize the objects
    web_scraper_obj = WebScraper()

    # sheet.max_row + 1
    for row in range(11, 12):
        startup_name = sheet.cell(row=row, column=1).value
        url = sheet.cell(row=row, column=2).value
        redirect_url = sheet.cell(row=row, column=3).value
        use_cases_combined = sheet.cell(row=row, column=9).value

        if pd.isnull(url):
            continue
        elif pd.isnull(use_cases_combined):
            use_cases_combined = ""

        print(f"Startup Name:{startup_name}")

        prompts_obj = Prompts(TOTAL_PAGE_CRAWLS)


        # Prompt based approach for the EU AI Act
        eu_ai_act_prompt = prepare_AI_Act_prompt(master_prompt, use_cases_combined)
        # print(f"EU AI Act Prompt: {eu_ai_act_prompt}\n\n\n")

        # eu_ai_act_obj = ChatGPT(classification_model_name, eu_ai_act_prompt, [], OpenAI(api_key=os.getenv("MY_KEY"), max_retries=5))
        # eu_ai_act_response, input_tokens, output_tokens = eu_ai_act_obj.chat_model()
        # # Update token cost
        # web_scraper_obj.set_token_cost(input_tokens, output_tokens, classification_model_name)


        # use anthropic to clasify
        eu_ai_act_response, input_tokens, output_tokens = claude_api(eu_ai_act_prompt)
        # Update token cost
        web_scraper_obj.set_token_cost(input_tokens, output_tokens, "claude-3-7-sonnet-20250219")


        save_to_excel(output_sheet, output_wb, startup_name, url, redirect_url, use_cases_combined, eu_ai_act_response, web_scraper_obj.get_token_cost(), output_filename)

        # --- Finishing calls ---
        # Reset token cost, redirected URL
        web_scraper_obj.reset_token_cost()
        web_scraper_obj.reset_redirect_url()








if __name__ == "__main__":

    startups_file = "Local Output/All Use Cases Combined.xlsx"
    sheet = load_startups_excel(startups_file)

    output_sheet, output_wb = create_results_file()

    output_filename = "Updated Claude Sonnet Results.xlsx"

    prompt_approach(classification_model_name='o3-mini', sheet=sheet, output_sheet=output_sheet, output_wb=output_wb, output_filename=output_filename)
    


