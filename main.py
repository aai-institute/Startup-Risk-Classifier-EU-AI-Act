# Standard Library
import os
import re
import ast
import json
import time
import csv

# Third-Party Library
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI
import anthropic
from openpyxl import Workbook

# Local Imports
from Classes import ChatGPT, Prompts, WebScraper, TextExtractor
from large_prompts.master_prompt import master_prompt
from re_functions.use_case_extractor import extract_use_cases
from simple_model_functions.api_functions import gemini_api, mistral_api


# Load environment variables
load_dotenv()

# Constants
TOTAL_PAGE_CRAWLS = 4


def claude_api(model, prompt):
    try:
        client = anthropic.Anthropic(
            api_key=os.getenv("MY_ANTHROPIC_KEY")
        )
        message = client.messages.create(
            model=model,
            max_tokens=8192,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        message_content = message.content[0].text  # Extracts the actual message text
        input_tokens = message.usage.input_tokens  # Extracts input tokens
        output_tokens = message.usage.output_tokens  # Extracts output tokens

    except Exception as e:
        message_content = f"API ERROR: Anthropic API failed"
        input_tokens = 0
        output_tokens = 0

    return message_content, input_tokens, output_tokens

def load_startups_excel(startups_file):
    sheet = openpyxl.load_workbook(startups_file)["AI Use Cases"]
    return sheet

def create_results_file():    
    # Save the results to a new Excel file
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "AI Use Cases"

    return output_sheet, output_wb

def content_shortener(content_shortener_model, prompts_obj, page_content):
    chat_shorten_page_obj = ChatGPT(content_shortener_model, prompts_obj.shorten_page_content(page_content), [], OpenAI(api_key=os.getenv("MY_1_KEY"), max_retries=5))
    chat_shorten_page_response, input_tokens, output_tokens = chat_shorten_page_obj.chat_model()
    # print(f"Shortened Page Content: {chat_shorten_page_response}")

    return chat_shorten_page_response, input_tokens, output_tokens

def traverse_links(web_scraper_obj, links, model_name, content_shortener_model, ai_use_cases, prompts_obj):
    
    try:
        # Traverse the important links
        for link in links:
            web_scraper_obj.set_url(link)
            # print(f"Going into relavant link: {link}")
            web_scraper_obj.load_page()
            page_content = web_scraper_obj.get_page_content(model_name)

            # Shorten the content
            shortened_content, input_tokens, output_tokens = content_shortener(content_shortener_model, prompts_obj, page_content)
            # Update token cost
            web_scraper_obj.set_token_cost(input_tokens, output_tokens, content_shortener_model)


            chat_use_case_obj = ChatGPT(model_name, prompts_obj.update_startup_summary(f"\n\n".join(ai_use_cases), shortened_content), [], OpenAI(api_key=os.getenv("MY_1_KEY"), max_retries=5))
            chat_use_case_response, input_tokens, output_tokens = chat_use_case_obj.chat_model()

            # Update token cost
            web_scraper_obj.set_token_cost(input_tokens, output_tokens, model_name)
            
            ai_use_cases.append(chat_use_case_response)
    
    except Exception as e:
        # Return error message only
        return ["Page Error: Unable to traverse links"]


    return ai_use_cases

def extract_list(input_string):
    # Regular expression to match Python-style lists
    list_pattern = r"\[.*?\]"
    match = re.search(list_pattern, input_string)
    if match:
        return ast.literal_eval(match.group())
    return None


def save_to_excel(output_sheet, output_wb, startup_name, url, redirect_url, use_cases_combined, eu_ai_act_response, total_token_cost, output_filename):
    headers = ["Startup Name", "Homepage URL", "Redirected URL", "AI Use Cases" , "EU AI Act Risk Classification", "Total Token Cost ($)"]
    
    # Write headers if not present
    if output_sheet.max_row < 2:
        output_sheet.append(headers)

    # Write data
    
    row = [startup_name, url, redirect_url, use_cases_combined, eu_ai_act_response, total_token_cost]

    output_sheet.append(row)
    output_wb.save(f"{output_filename}")


def extract_use_cases_from_response(use_cases_full_text):
    all_use_cases = []
    # Extract each AI use case from the full text
    text_extractor = TextExtractor(use_cases_full_text)
    text_extractor.set_use_cases()
    df_use_cases = text_extractor.get_use_cases()
    
    for index, row in df_use_cases.iterrows():
        row_string = "\n".join(f"{col}: {val}" for col, val in row.items())
        all_use_cases.append(row_string + '\n\n')
    
    return all_use_cases    


def gpt_search(web_search_model, url, prompts_obj, web_scraper_obj):
    gpt_use_case_response = ""
    for _ in range(6):
        # Use the web search tool to generate new use cases
        gpt_use_case_obj = ChatGPT(web_search_model, prompts_obj.generate_use_case_gpt(url), [], OpenAI(api_key=os.getenv("MY_1_KEY"), max_retries=5))
        gpt_use_case_response, input_tokens, output_tokens = gpt_use_case_obj.chat_model()
        # Update token cost
        web_scraper_obj.set_token_cost(input_tokens, output_tokens, web_search_model)

        if gpt_use_case_response and "No information found" not in gpt_use_case_response:
            break
        else:
            print(f"\nNo Use Cases generated.Retrying...\n")
            gpt_use_case_response = ""

    return gpt_use_case_response


def call_model_with_retry(model_name, model_type, formatted_prompt, web_scraper_obj, max_retries=4, retry_delay=10):
    """
    Generic function to call any model with retry logic
    """
    for attempt in range(max_retries):
        try:
            if model_type == "chatgpt":
                chat_obj = ChatGPT(model_name, formatted_prompt, [], OpenAI(api_key=os.getenv("MY_1_KEY"), max_retries=5))
                response, input_tokens, output_tokens = chat_obj.chat_model()
            elif model_type == "claude":
                response, input_tokens, output_tokens = claude_api(model_name, formatted_prompt)
            elif model_type == "deepseek":
                chat_obj = ChatGPT(model_name, formatted_prompt, [], OpenAI(api_key=os.getenv("DEEPSEEK_KEY"), max_retries=5, base_url="https://api.deepseek.com"))
                response, input_tokens, output_tokens = chat_obj.chat_model()
            elif model_type == "gemini":
                response, input_tokens, output_tokens = gemini_api(model_name, formatted_prompt)
            elif model_type == "mistral":
                response, input_tokens, output_tokens = mistral_api(model_name, formatted_prompt)
            else:
                raise ValueError(f"Unknown model type: {model_type}")
            
            # Update token cost
            web_scraper_obj.set_token_cost(input_tokens, output_tokens, model_name)
            response += "\n\n########END OF USE CASE########\n\n"
            return response
            
        except Exception as e:
            print(f"Error in {model_type.upper()} API call. Attempt {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
            else:
                web_scraper_obj.set_token_cost(input_tokens=0, output_tokens=0, model_name=model_name)
                raise RuntimeError(f"{model_type.upper()} Classification failed") from e


def multiple_model_approach(chatgpt_model, claude_model, deepseek_model, gemini_model, mistral_model):
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    workbook_filename = "Results.xlsx" # File to store the results
    # Add a header row
    ws.append(["Startup Name", "Generated Text", "Token Cost ($)"])
    wb.save(workbook_filename)
    # Excel start row number
    row_num = 2


    # CSV
    csv_filename = "Results.csv" # Store in a separate CSV file
    with open(csv_filename, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Startup Name", "Generated Text", "Token Cost ($)"])


    # Allowed categories
    allowed_categories = [
        'Prohibited AI system',
        'High-risk AI system under Annex I',
        'High-risk AI system under Annex III',
        'High-risk AI system with transparency obligations',
        'System with transparency obligations',
        'Low-risk AI system',
        'Uncertain'
    ]

    # Initialize the objects
    web_scraper_obj = WebScraper()
    prompts_obj = Prompts(TOTAL_PAGE_CRAWLS)

    MAX_API_TRIES = 4
    retry_delay = 10

    start_index = 0
    end_index = 200

    with open('datasets/Use Cases/example_use_cases.json', 'r') as file:
        json_data = json.load(file)

        companies = json_data['companies']
        for idx in range(start_index, end_index):
            if idx < len(companies):
                company = companies[idx]
                startup_name = company['company_name']
                print(f"JSON Index: {idx}\nStartup Name: {startup_name}")

                # Define model configurations
                model_configs = [
                    (chatgpt_model, "chatgpt", "ChatGPT 4o"),
                    (claude_model, "claude", "Claude 3.7 Sonnet"),
                    (deepseek_model, "deepseek", "DeepSeek Reasoner"),
                    (gemini_model, "gemini", "Gemini 2.0 Flash Thinker"),
                    (mistral_model, "mistral", "Mistral Large")
                ]

                longest_reasoned_use_case_string = ""
                for use_case in company['use_cases']:
                    # print(use_case)
                    use_case_string = f"AI Use Case: {use_case["use_case_name"]}\nUse Case Description: {use_case["use_case_description"]}"
                    # print(use_case_string)

                    # Prepare the prompt with the use case
                    formatted_prompt = prompts_obj.prepare_AI_Act_prompt(master_prompt, use_case_string)
                    
                    # Call all models and collect responses
                    model_responses = []
                    for model_name, model_type, display_name in model_configs:
                        response = call_model_with_retry(model_name, model_type, formatted_prompt, web_scraper_obj, MAX_API_TRIES, retry_delay)
                        model_responses.append(response)
                        # print(f"\n\nFrom {display_name}:\n{response}")

                    # Combine all responses
                    final_string = "".join(model_responses)


                    # Get the combined json from all models
                    result_json = extract_use_cases(use_case, final_string)
                    # print(f"Result JSON: {result_json}")

                    with open('test.json', 'w') as json_file:
                        json.dump(result_json, json_file, indent=4)
                    

                    # Create a dictionary to store the votes and store individual classifications from all models
                    # Use display names from model configuration
                    voters = [config[2] for config in model_configs]
                    votings = {}
                    classifications_list = []
                    # Iterate through the result JSON and count votes for each classification
                    for model_use_case in result_json:
                        classification = model_use_case["Risk Classification"]
                        if classification in allowed_categories:
                            if classification not in votings:
                                votings[classification] = 0
                            votings[classification] += 1
                            classifications_list.append(classification)
                        else:
                            # If the classification is not in the allowed categories, re-classify it as "Uncertain"
                            model_use_case["Risk Classification"] = "Uncertain"
                            classifications_list.append("Uncertain")
                            if "Uncertain" not in votings:
                                votings["Uncertain"] = 0
                            votings["Uncertain"] += 1


                        
                    # Find the classification with the most votes.
                    max_votes = max(votings.values())
                    classifications_with_max_votes = [classification for classification, votes in votings.items() if votes == max_votes]
                    # 3. If tie, pick least risky from the tie group
                    if len(classifications_with_max_votes) > 1:
                        classifications_with_max_votes.sort(
                            key=lambda x: allowed_categories.index(x) if x in allowed_categories else -1,
                            reverse=True
                        )
                    final_classification = classifications_with_max_votes[0]

                    print(f"Votings: {votings}")
                    print(f"Max Votes: {max_votes}")
                    print(f"Classifications with max votes: {classifications_with_max_votes}")
                    
                    print(f"Classifications List: {classifications_list}")
                    print(f"Final Classification: {final_classification}")


                    # ***STORE THE VOTE DISTRIBUTION FROM EACH MODEL***
                    model_distribution = dict(zip(voters, classifications_list))
                    model_distribution_string = ""
                    for model, classification in model_distribution.items():
                        # print(f"{model}: {classification}")
                        model_distribution_string += f"{model}: {classification}\n"


                    # Get the use cases with the final classification
                    filtered_use_cases = [model_use_case for model_use_case in result_json if model_use_case["Risk Classification"] == final_classification]
                    # Get the use case with the longest reason
                    longest_reasoned_use_case = max(filtered_use_cases, key=lambda x: len(x["Reason"]))
                    longest_reasoned_use_case_index = result_json.index(longest_reasoned_use_case)


                    # Attach the distribution to the longest reasoned use case
                    longest_reasoned_use_case["Model Distribution"] = "\n"
                    for model_distrib_index, (k,v) in enumerate(model_distribution.items()):
                        longest_reasoned_use_case["Model Distribution"] += f"{k} => {v}"
                        longest_reasoned_use_case["Model Distribution"] += "\n" if model_distrib_index < len(model_distribution) - 1 else ""
                    # Attach which model it was classified from
                    longest_reasoned_use_case["Chosen Model"] = voters[longest_reasoned_use_case_index]

                    # print(f"Longest Reasoned Use Case: {longest_reasoned_use_case}")
                    
                    
                    
                    longest_reasoned_use_case_string += "\n".join(f"{k}: {v}" for k, v in longest_reasoned_use_case.items())
                    longest_reasoned_use_case_string += "\n\n########END OF CLASSIFICATION########\n\n\n\n"

                    # print(f"Model Distribution:\n{model_distribution_string}")
                    print(f"{longest_reasoned_use_case_string}")

                    # break


            # Write the final use case string to Excel
            ws.cell(row=row_num, column=1, value=startup_name)
            ws.cell(row=row_num, column=2, value=longest_reasoned_use_case_string)
            ws.cell(row=row_num, column=3, value=web_scraper_obj.get_token_cost())
            wb.save(workbook_filename)
            
            # Also write the same row to CSV
            with open(csv_filename, mode="a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([startup_name, longest_reasoned_use_case_string, web_scraper_obj.get_token_cost()])
            
            row_num += 1

            # reset token cost, redirected URL
            web_scraper_obj.reset_token_cost()
            web_scraper_obj.reset_redirect_url()


            # break


if __name__ == "__main__":
    multiple_model_approach(chatgpt_model="chatgpt-4o-latest", claude_model="claude-3-7-sonnet-20250219", deepseek_model="deepseek-reasoner", gemini_model="gemini-2.0-flash-thinking-exp-01-21", mistral_model="mistral-large-latest")
    


